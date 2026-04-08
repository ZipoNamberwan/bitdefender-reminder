from pathlib import Path

from openpyxl import load_workbook


REFERENSI_SHEET = 'Refrensi'
START_SATKER = 3501
END_SATKER = 3579
START_ROW_SATKER = 3


def first_four_chars(value):
	if value is None:
		return ''
	return str(value).strip()[:4]


def in_satker_range(prefix):
	if len(prefix) != 4 or not prefix.isdigit():
		return False

	num = int(prefix)
	return START_SATKER <= num <= END_SATKER


def build_header_index(sheet):
	header_row = sheet[1]
	header_index = {}

	for idx, cell in enumerate(header_row, start=1):
		value = cell.value
		if value is None:
			continue
		key = str(value).strip()
		if key:
			header_index[key] = idx

	return header_index


def get_required_columns(header_index):
	tag_column = None
	if 'Satker' in header_index:
		tag_column = 'Satker'
	elif 'Tag Satker' in header_index:
		tag_column = 'Tag Satker'

	required = ['Device Name', 'Users', 'NUP', 'Update']
	missing = [col for col in required if col not in header_index]
	if tag_column is None:
		missing.append('Tag Satker/Satker')

	if missing:
		missing_text = ', '.join(missing)
		raise ValueError(f'Missing required columns in referensi sheet: {missing_text}')

	col_idx = {name: header_index[name] for name in required}
	col_idx['Satker Match'] = header_index[tag_column]
	return col_idx


def read_referensi_rows(sheet, col_idx):
	grouped = {}

	for row in range(2, sheet.max_row + 1):
		tag_satker = sheet.cell(row=row, column=col_idx['Satker Match']).value
		if tag_satker is None:
			continue

		tag_key = first_four_chars(tag_satker)
		if not tag_key:
			continue

		grouped.setdefault(tag_key, []).append(
			{
				'Satker': str(tag_satker).strip(),
				'Device Name': sheet.cell(row=row, column=col_idx['Device Name']).value,
				'User': sheet.cell(row=row, column=col_idx['Users']).value,
				'NUP': sheet.cell(row=row, column=col_idx['NUP']).value,
				'Update': sheet.cell(row=row, column=col_idx['Update']).value,
			}
		)

	return grouped


def clear_target_columns(sheet, from_row):
	for row in range(from_row, sheet.max_row + 1):
		sheet.cell(row=row, column=1).value = None   # A -> No
		sheet.cell(row=row, column=2).value = None   # B -> Satker
		sheet.cell(row=row, column=6).value = None   # F -> NUP
		sheet.cell(row=row, column=11).value = None  # K -> Device Name
		sheet.cell(row=row, column=12).value = None  # L -> Update
		sheet.cell(row=row, column=13).value = None  # M -> User


def fill_satker_sheet(sheet, rows):
	clear_target_columns(sheet, START_ROW_SATKER)

	target_row = START_ROW_SATKER
	for no, item in enumerate(rows, start=1):
		sheet.cell(row=target_row, column=1).value = no                        # A
		sheet.cell(row=target_row, column=2).value = item.get('Satker')        # B
		sheet.cell(row=target_row, column=11).value = item.get('Device Name')  # K
		sheet.cell(row=target_row, column=13).value = item.get('User')         # M
		sheet.cell(row=target_row, column=6).value = item.get('NUP')           # F
		sheet.cell(row=target_row, column=12).value = item.get('Update')       # L
		target_row += 1


def transform_file(input_path, output_path):
	workbook = load_workbook(input_path)

	if REFERENSI_SHEET not in workbook.sheetnames:
		raise ValueError(f'Sheet {REFERENSI_SHEET!r} not found in workbook')

	referensi = workbook[REFERENSI_SHEET]
	header_index = build_header_index(referensi)
	col_idx = get_required_columns(header_index)
	grouped_rows = read_referensi_rows(referensi, col_idx)

	for satker_name in workbook.sheetnames:
		if satker_name == REFERENSI_SHEET:
			continue

		satker_prefix = first_four_chars(satker_name)
		if not in_satker_range(satker_prefix):
			continue

		satker_rows = grouped_rows.get(satker_prefix, [])
		fill_satker_sheet(workbook[satker_name], satker_rows)
		print(f'Sheet {satker_name} ({satker_prefix}): wrote {len(satker_rows)} rows')

	workbook.save(output_path)
	print(f'Saved output file: {output_path}')


def main():
	base_dir = Path(__file__).resolve().parent
	input_path = base_dir / 'input.xlsx'
	output_path = base_dir / 'input_filled.xlsx'

	transform_file(input_path, output_path)


if __name__ == '__main__':
	main()
