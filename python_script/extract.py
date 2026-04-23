from pathlib import Path

from openpyxl import load_workbook


REFERENSI_SHEET = 'Refrensi'
START_SATKER = 3501
END_SATKER = 3579
START_ROW_SATKER = 3
START_ROW_REFERENSI = 2


def first_four_chars(value):
	if value is None:
		return ''
	return str(value).strip()[:4]


def in_satker_range(prefix):
	if len(prefix) != 4 or not prefix.isdigit():
		return False

	num = int(prefix)
	return START_SATKER <= num <= END_SATKER


def build_header_index(sheet):
	header_row = sheet[1]
	header_index = {}

	for idx, cell in enumerate(header_row, start=1):
		value = cell.value
		if value is None:
			continue

		key = str(value).strip()
		if key:
			header_index[key] = idx

	return header_index


def get_required_columns(header_index):
	tag_column = None
	if 'Satker' in header_index:
		tag_column = 'Satker'
	elif 'Tag Satker' in header_index:
		tag_column = 'Tag Satker'

	required = ['Device Name', 'Users', 'NUP', 'Update']
	missing = [col for col in required if col not in header_index]
	if tag_column is None:
		missing.append('Tag Satker/Satker')

	if missing:
		missing_text = ', '.join(missing)
		raise ValueError(f'Missing required columns in referensi sheet: {missing_text}')

	col_idx = {name: header_index[name] for name in required}
	col_idx['Satker Match'] = header_index[tag_column]
	return col_idx


def has_satker_row_data(sheet, row):
	# Satker sheet payload columns from transform.py mapping.
	values = [
		sheet.cell(row=row, column=2).value,   # B -> Satker
		sheet.cell(row=row, column=11).value,  # K -> Device Name
		sheet.cell(row=row, column=13).value,  # M -> User
		sheet.cell(row=row, column=6).value,   # F -> NUP
		sheet.cell(row=row, column=12).value,  # L -> Update
	]
	return any(value not in (None, '') for value in values)


def read_satker_rows(sheet):
	rows = []

	for row in range(START_ROW_SATKER, sheet.max_row + 1):
		if not has_satker_row_data(sheet, row):
			continue

		satker = sheet.cell(row=row, column=2).value
		device_name = sheet.cell(row=row, column=11).value
		user_value = sheet.cell(row=row, column=13).value
		nup = sheet.cell(row=row, column=6).value
		update = sheet.cell(row=row, column=12).value

		rows.append(
			{
				'Satker': str(satker).strip() if satker is not None else '',
				'Device Name': device_name,
				'Users': user_value,
				'NUP': nup,
				'Update': update,
			}
		)

	return rows


def clear_referensi_rows(sheet, col_idx, from_row):
	for row in range(from_row, sheet.max_row + 1):
		sheet.cell(row=row, column=col_idx['Satker Match']).value = None
		sheet.cell(row=row, column=col_idx['Device Name']).value = None
		sheet.cell(row=row, column=col_idx['Users']).value = None
		sheet.cell(row=row, column=col_idx['NUP']).value = None
		sheet.cell(row=row, column=col_idx['Update']).value = None


def write_referensi_rows(sheet, col_idx, rows):
	clear_referensi_rows(sheet, col_idx, START_ROW_REFERENSI)

	target_row = START_ROW_REFERENSI
	for item in rows:
		sheet.cell(row=target_row, column=col_idx['Satker Match']).value = item.get('Satker')
		sheet.cell(row=target_row, column=col_idx['Device Name']).value = item.get('Device Name')
		sheet.cell(row=target_row, column=col_idx['Users']).value = item.get('Users')
		sheet.cell(row=target_row, column=col_idx['NUP']).value = item.get('NUP')
		sheet.cell(row=target_row, column=col_idx['Update']).value = item.get('Update')
		target_row += 1


def extract_file(input_path, output_path):
	workbook = load_workbook(input_path)

	if REFERENSI_SHEET not in workbook.sheetnames:
		raise ValueError(f'Sheet {REFERENSI_SHEET!r} not found in workbook')

	referensi = workbook[REFERENSI_SHEET]
	header_index = build_header_index(referensi)
	col_idx = get_required_columns(header_index)

	extracted_rows = []
	for satker_name in workbook.sheetnames:
		if satker_name == REFERENSI_SHEET:
			continue

		satker_prefix = first_four_chars(satker_name)
		if not in_satker_range(satker_prefix):
			continue

		satker_rows = read_satker_rows(workbook[satker_name])
		extracted_rows.extend(satker_rows)
		print(f'Sheet {satker_name} ({satker_prefix}): read {len(satker_rows)} rows')

	write_referensi_rows(referensi, col_idx, extracted_rows)

	workbook.save(output_path)
	print(f'Saved output file: {output_path}')
	print(f'Total extracted rows: {len(extracted_rows)}')


def main():
	base_dir = Path(__file__).resolve().parent
	input_path = base_dir / 'input_filled.xlsx'
	output_path = base_dir / 'input_extracted.xlsx'

	extract_file(input_path, output_path)


if __name__ == '__main__':
	main()
